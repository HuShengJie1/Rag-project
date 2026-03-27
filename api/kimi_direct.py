import json
import os
import shutil
import uuid
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from dotenv import load_dotenv
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from langchain_core.messages import AIMessage, HumanMessage, SystemMessage
from openai import OpenAI
from pydantic import BaseModel

from core import llm

load_dotenv()

router = APIRouter(prefix="/api/kimi")

PROJECT_ROOT = Path(__file__).resolve().parent.parent
UPLOAD_DIR = PROJECT_ROOT / "data" / "kimi_uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def _get_kimi_client() -> OpenAI:
    api_key = os.getenv("KIMI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="KIMI_API_KEY 未配置")
    base_url = os.getenv("KIMI_BASE_URL", "https://api.moonshot.cn/v1")
    return OpenAI(api_key=api_key, base_url=base_url)


def _save_upload(file: UploadFile, upload_id: str) -> Path:
    suffix = Path(file.filename).suffix if file.filename else ""
    save_path = UPLOAD_DIR / f"{upload_id}{suffix}"
    with save_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return save_path


def _extract_text_with_kimi(file_path: Path) -> str:
    client = _get_kimi_client()
    file_object = client.files.create(file=file_path, purpose="file-extract")
    file_content = client.files.content(file_id=file_object.id).text
    return file_content or ""


def _trim_table(rows: List[List[str]]) -> List[List[str]]:
    if not rows:
        return rows
    last_nonempty_row = -1
    last_nonempty_col = -1
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            if str(val).strip() != "":
                last_nonempty_row = max(last_nonempty_row, r_idx)
                last_nonempty_col = max(last_nonempty_col, c_idx)
    if last_nonempty_row == -1 or last_nonempty_col == -1:
        return []
    trimmed = []
    for r in rows[: last_nonempty_row + 1]:
        trimmed.append(r[: last_nonempty_col + 1])
    return trimmed


def _extract_excel_tables(file_path: Path) -> Dict:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，无法解析 Excel") from e

    wb = load_workbook(file_path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows: List[List[str]] = []
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
        rows = _trim_table(rows)
        sheets.append({"name": ws.title, "rows": rows})
    return {"sheets": sheets}


def _persist_upload_meta(upload_id: str, filename: str, has_table: bool) -> None:
    meta_path = UPLOAD_DIR / f"{upload_id}.meta.json"
    meta_path.write_text(json.dumps({
        "id": upload_id,
        "name": filename,
        "has_table": has_table
    }, ensure_ascii=False), encoding="utf-8")


@router.post("/upload")
async def kimi_upload(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名为空")

    upload_id = uuid.uuid4().hex
    raw_path = _save_upload(file, upload_id)
    suffix = raw_path.suffix.lower()
    has_table = suffix in [".xlsx", ".xls"]

    try:
        text = _extract_text_with_kimi(raw_path)
        text_path = UPLOAD_DIR / f"{upload_id}.txt"
        text_path.write_text(text, encoding="utf-8")
        if has_table:
            table_data = _extract_excel_tables(raw_path)
            table_path = UPLOAD_DIR / f"{upload_id}.table.json"
            table_path.write_text(json.dumps(table_data, ensure_ascii=False), encoding="utf-8")
    finally:
        # 原始文件可按需保留；默认删除以节省空间
        try:
            raw_path.unlink(missing_ok=True)
        except Exception:
            pass

    _persist_upload_meta(upload_id, file.filename, has_table)
    return {"id": upload_id, "name": file.filename, "has_table": has_table}


class KimiChatRequest(BaseModel):
    prompt: str
    upload_id: Optional[str] = None
    history: List[Dict[str, str]] = []


@router.post("/chat")
async def kimi_chat(request: KimiChatRequest):
    messages = [
        SystemMessage(content="你是一个严谨的文档助手，请基于提供的文件内容回答问题。")
    ]

    if request.upload_id:
        text_path = UPLOAD_DIR / f"{request.upload_id}.txt"
        if not text_path.exists():
            raise HTTPException(status_code=404, detail="文件内容未找到")
        file_text = text_path.read_text(encoding="utf-8")
        if file_text.strip():
            messages.append(SystemMessage(content=file_text))

    clean_history = request.history[-6:] if request.history else []
    for msg in clean_history:
        role = msg.get("role")
        content = msg.get("content", "")
        if not content:
            continue
        if role == "user":
            messages.append(HumanMessage(content=content))
        elif role == "assistant":
            messages.append(AIMessage(content=content))

    messages.append(HumanMessage(content=request.prompt))

    async def stream_generator():
        yield json.dumps({"evidence": []}, ensure_ascii=False) + "---METADATA_SEPARATOR---"
        async for chunk in llm.astream(messages):
            if chunk.content:
                yield chunk.content

    return StreamingResponse(stream_generator(), media_type="text/event-stream")


@router.get("/table/{upload_id}")
async def get_kimi_table(upload_id: str):
    table_path = UPLOAD_DIR / f"{upload_id}.table.json"
    if not table_path.exists():
        raise HTTPException(status_code=404, detail="表格内容未找到")
    data = json.loads(table_path.read_text(encoding="utf-8"))
    return JSONResponse(content=data)
